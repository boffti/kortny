#!/usr/bin/env python3
"""Build a styled .xlsx workbook from a JSON spec — runs in the Kortny sandbox.

The spec describes one or more sheets. Each sheet carries a header row,
data rows, optional column number-formats, and optional formula columns.
The script applies the professional conventions Kortny commits to:

  * frozen header row (always row 1), bold header with fill
  * per-column Excel number formats (currency / percent / thousands / date)
  * financial color-coding: negative numbers render red, positive green
    on any column flagged ``"signed": true``
  * formula columns are written as real Excel formulas (``=B2*C2``) so the
    workbook computes live — never as pre-baked strings

Network is never required; everything is local file I/O. Input is a JSON
file (``--spec``) or stdin; output is a path under the workbench dir.

Spec shape (see references/spec-format.md for the full reference):

    {
      "sheets": [
        {
          "name": "Q1 Forecast",
          "columns": [
            {"header": "Region", "key": "region"},
            {"header": "Revenue", "key": "rev", "format": "currency", "signed": true},
            {"header": "Growth", "key": "growth", "format": "percent", "signed": true}
          ],
          "rows": [
            {"region": "EMEA", "rev": 120000, "growth": 0.12},
            {"region": "APAC", "rev": -5000, "growth": -0.03}
          ],
          "formula_columns": [
            {"header": "Rev x2", "formula": "={rev}{row}*2", "format": "currency"}
          ],
          "total_row": true
        }
      ]
    }
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# Excel number-format strings keyed by the friendly names accepted in specs.
# The currency/percent formats include the red-negative pattern so Excel itself
# colors losses red regardless of cell font — belt-and-suspenders with the
# explicit signed-cell coloring below.
NUMBER_FORMATS: dict[str, str] = {
    "currency": "#,##0.00;[Red]-#,##0.00",
    "currency_whole": "#,##0;[Red]-#,##0",
    "percent": "0.0%;[Red]-0.0%",
    "thousands": "#,##0;[Red]-#,##0",
    "number": "0.00;[Red]-0.00",
    "date": "yyyy-mm-dd",
    "text": "@",
}

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TOTAL_FONT = Font(bold=True)
POSITIVE_FONT = Font(color="166534")  # green-800
NEGATIVE_FONT = Font(color="B91C1C")  # red-700


def _resolve_format(name: str | None) -> str | None:
    if name is None:
        return None
    if name not in NUMBER_FORMATS:
        raise ValueError(
            f"unknown format {name!r}; valid: {', '.join(sorted(NUMBER_FORMATS))}"
        )
    return NUMBER_FORMATS[name]


def _apply_signed_color(cell: Any, value: Any) -> None:
    """Color a numeric cell green/red by sign (financial convention)."""
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        cell.font = NEGATIVE_FONT if value < 0 else POSITIVE_FONT


def _write_sheet(ws: Worksheet, sheet: dict[str, Any]) -> None:
    columns: list[dict[str, Any]] = list(sheet.get("columns", []))
    formula_columns: list[dict[str, Any]] = list(sheet.get("formula_columns", []))
    rows: list[dict[str, Any]] = list(sheet.get("rows", []))

    all_headers = [c["header"] for c in columns] + [
        c["header"] for c in formula_columns
    ]
    if not all_headers:
        raise ValueError(f"sheet {ws.title!r} has no columns")

    # Header row.
    for col_idx, header in enumerate(all_headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Map each data column key to its letter for formula substitution.
    key_to_letter: dict[str, str] = {}
    for idx, col in enumerate(columns, start=1):
        key = col.get("key")
        if key:
            key_to_letter[key] = get_column_letter(idx)

    # Data rows.
    for r_offset, row in enumerate(rows):
        excel_row = r_offset + 2  # row 1 is the header
        for col_idx, col in enumerate(columns, start=1):
            value = row.get(col["key"])
            cell = ws.cell(row=excel_row, column=col_idx, value=value)
            fmt = _resolve_format(col.get("format"))
            if fmt:
                cell.number_format = fmt
            if col.get("signed"):
                _apply_signed_color(cell, value)
        # Formula columns reference data columns by {key}{row}.
        for f_offset, fcol in enumerate(formula_columns):
            col_idx = len(columns) + 1 + f_offset
            template: str = fcol["formula"]
            formula = template.replace("{row}", str(excel_row))
            for key, letter in key_to_letter.items():
                formula = formula.replace("{" + key + "}", letter)
            cell = ws.cell(row=excel_row, column=col_idx, value=formula)
            fmt = _resolve_format(fcol.get("format"))
            if fmt:
                cell.number_format = fmt

    # Optional total row summing every numeric (signed/format) column.
    if sheet.get("total_row") and rows:
        total_row_idx = len(rows) + 2
        first = ws.cell(row=total_row_idx, column=1, value="Total")
        first.font = TOTAL_FONT
        last_data_row = len(rows) + 1
        for col_idx, col in enumerate(columns, start=1):
            if col_idx == 1:
                continue
            if col.get("format") in {
                "currency",
                "currency_whole",
                "thousands",
                "number",
            }:
                letter = get_column_letter(col_idx)
                cell = ws.cell(
                    row=total_row_idx,
                    column=col_idx,
                    value=f"=SUM({letter}2:{letter}{last_data_row})",
                )
                cell.font = TOTAL_FONT
                fmt = _resolve_format(col.get("format"))
                if fmt:
                    cell.number_format = fmt

    # Freeze the header row and widen columns to their longest cell.
    ws.freeze_panes = "A2"
    for col_idx, header in enumerate(all_headers, start=1):
        letter = get_column_letter(col_idx)
        longest = len(str(header))
        for r in range(2, ws.max_row + 1):
            val = ws.cell(row=r, column=col_idx).value
            if val is not None:
                longest = max(longest, len(str(val)))
        ws.column_dimensions[letter].width = min(max(longest + 2, 10), 48)


def build(spec: dict[str, Any], out_path: Path) -> Path:
    sheets = spec.get("sheets")
    if not sheets:
        raise ValueError("spec must contain at least one sheet under 'sheets'")
    wb = Workbook()
    # Workbook() ships one default sheet; reuse it for the first spec sheet.
    default_ws = wb.active
    for idx, sheet in enumerate(sheets):
        if idx == 0 and default_ws is not None:
            ws = default_ws
            ws.title = sheet.get("name", "Sheet1")[:31]
        else:
            ws = wb.create_sheet(title=sheet.get("name", f"Sheet{idx + 1}")[:31])
        _write_sheet(ws, sheet)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Build a styled .xlsx workbook.")
    parser.add_argument(
        "--spec",
        help="path to a JSON spec file; omit to read JSON from stdin",
    )
    parser.add_argument(
        "--out",
        required=True,
        help="output .xlsx path (under the workbench dir)",
    )
    args = parser.parse_args(argv)

    raw = Path(args.spec).read_text() if args.spec else sys.stdin.read()
    spec = json.loads(raw)
    out = build(spec, Path(args.out))
    print(f"wrote workbook: {out} ({out.stat().st_size} bytes)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
